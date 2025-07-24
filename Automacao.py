import requests
import json
import pandas as pd
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles.differential import DifferentialStyle
TOKEN = "eyJhbGciOiJIUzUxMiJ9.eyJpc3MiOiJQaXBlZnkiLCJpYXQiOjE3NTA0MTk3MTEsImp0aSI6IjkxNTU3NjM4LTAwZmMtNDJlYS05NDRjLWE5NWI3MTk2MGMxZCIsInN1YiI6MzA2NzM5NzE4LCJ1c2VyIjp7ImlkIjozMDY3Mzk3MTgsImVtYWlsIjoiZGVzZW52b2x2aW1lbnRvMkB0NXRlYy5jb20uYnIifX0.nTiOpnKPxH_E0YcMih3leVJUuaW9Fr6cq6Vr_jJ2IZeT61J2-0desMbdKW8O8A2Z_ta6Cgt1BtohpIvtj_x-1Q"  # Obtenha em: Pipefy > Perfil > API Tokens
TABLE_ID = "304107875"
def colocarMascara_cidade(cidade):
    # Se for lista, pega o primeiro elemento
    if isinstance(cidade, list) and cidade:
        cidade = cidade[0]
    # Se for string com colchetes, remove colchetes e aspas
    if isinstance(cidade, str):
        cidade = cidade.strip().replace('[', '').replace(']', '').replace('"', '')
    return cidade.strip()
def colocarMascara_numero(numero):
    if not numero:
        return ""
    numero = str(numero).strip().replace('.', '').replace('-', '')
    if len(numero) == 11:
        return f"({numero[:2]}) {numero[2:7]}-{numero[7:]}"

    elif len(numero) == 10:
        return f"({numero[:2]}) {numero[2:6]}-{numero[6:]}"

    elif len(numero) == 9:
        return f"{numero[:5]}-{numero[5:]}"

    elif len(numero) == 8:
        return f"{numero[:4]}-{numero[4:]}"
   
        
    return numero

def colocarMascara_cnpj(cnpj):
    cnpj = cnpj.strip().replace('.', '').replace('/', '').replace('-', '')
    if len(cnpj) == 14:
        return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
    else:
        return cnpj  # Retorna sem máscara se não tiver 14 dígitos
def consulta_numero_por_cnpj(cnpj):
    if(cnpj is None or cnpj.strip() == ''):
        return None
    cnpj_limpo = cnpj.strip().replace('.', '').replace('/', '').replace('-', '')
    url = f"https://minhareceita.org/{cnpj_limpo}"
    try:
        response = requests.get(url)
        response.raise_for_status()
        data = response.json()
        # O campo de telefone geralmente vem como "telefone" ou "telefone1"
        numero = colocarMascara_numero(data.get("ddd_telefone_1", ""))

        if data.get("ddd_telefone_2"):
            numero += ", " + colocarMascara_numero(data.get("ddd_telefone_2", ""))
        
        return numero
    except requests.RequestException as e:
        print(f"Erro ao consultar CNPJ {cnpj}: {e}")
        return None
    
def consulta_socios_por_cnpj(cnpj):
    cnpj_limpo = cnpj.strip().replace('.', '').replace('/', '').replace('-', '')
    url = f"https://minhareceita.org/{cnpj_limpo}"
    try:
        response = requests.get(url)
        response.raise_for_status()
        data = response.json()
        socios = data.get("qsa", [])
        valor = ''
        if socios:
            if isinstance(socios, list):
                socios_formatados = []
                for socio in socios:
                    nome = socio.get('nome', '').strip()
                    qualificacao = socio.get('qualificacao', '').strip()
                    if nome:
                        if qualificacao:
                            socios_formatados.append(f"{nome} ({qualificacao})")
                        else:
                            socios_formatados.append(f"{nome}")
                valor = ', '.join(socios_formatados)
            else:
                valor = str(socios)
            if not valor:
                valor = 'Nenhum sócio encontrado'
        else:
            valor = 'Nenhum sócio encontrado'
        
        return valor
    except requests.RequestException as e:
        print(f"Erro ao consultar CNPJ {cnpj}: {e}")
        return []
def consulta_cnpj(cnpj, campos_desejados=None):
  
    cnpj_limpo = cnpj.strip().replace('.', '').replace('/', '').replace('-', '')
    url = f"https://minhareceita.org/{cnpj_limpo}"
    
    try:
        response = requests.get(url)
        response.raise_for_status()
        data = response.json()
        
        if campos_desejados:
            
            return {campo: data.get(campo) for campo in campos_desejados}
        print("data", data)
        return data
    except requests.RequestException as e:
        print(f"Erro ao consultar CNPJ {cnpj}: {e}")
        return None




def criar_excel_formatado(df, nome_arquivo="pipefy_records.xlsx"):
    """Cria um arquivo Excel formatado com estilos profissionais"""
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados Pipefy"
    
    # Adicionar dados do DataFrame
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Estilos para o cabeçalho
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    
    # Estilos para células de dados
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Bordas
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aplicar formatação ao cabeçalho
    for col in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Aplicar formatação às células de dados
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # Alternar cores das linhas (zebrado)
            if row % 2 == 0:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Ajustar largura das colunas automaticamente
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Máximo de 50 caracteres
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Congelar primeira linha (cabeçalho)
    ws.freeze_panes = 'A2'
    
    # Adicionar filtros automáticos
    ws.auto_filter.ref = ws.dimensions
    
    # Salvar arquivo
    wb.save(nome_arquivo)
    print(f"Arquivo Excel formatado salvo como {nome_arquivo}")

def criar_excel_formatado_com_ordenacao(df, nome_arquivo="pipefy_records_ordenado.xlsx", coluna_ordenacao=None, ordem_crescente=True):
    """Cria um arquivo Excel formatado e ordenado"""
    
    # Ordenar DataFrame se coluna de ordenacao for especificada
    if coluna_ordenacao and coluna_ordenacao in df.columns:
        df = df.sort_values(by=coluna_ordenacao, ascending=ordem_crescente)
    
    criar_excel_formatado(df, nome_arquivo)

def get_all_records():
    url = "https://api.pipefy.com/graphql"
    headers = {
        "Authorization": f"Bearer {TOKEN}",
        "Content-Type": "application/json"
    }
    
    all_records = []
    cursor = None
    page_count = 0
    
    while True:
        page_count += 1
        query = {
            "query": f"""
            {{
              table(id: "{TABLE_ID}") {{
                table_records(first: 100, after: {f'"{cursor}"' if cursor else "null"}) {{
                  pageInfo {{
                    hasNextPage
                    endCursor
                  }}
                  edges {{
                    node {{
                      id
                      title
                      created_at
                      record_fields {{
                        name
                        value
                      }}
                    }}
                  }}
                }}
              }}
            }}
            """
        }
        
        response = requests.post(url, json=query, headers=headers)
        data = response.json()
        
        if "errors" in data:
            print("Erro:", json.dumps(data["errors"], indent=2))
            break
            
        if "data" not in data:
            print("Resposta inesperada da API:", json.dumps(data, indent=2, ensure_ascii=False))
            break
        records_data = data["data"]["table"]["table_records"]
        records = [edge["node"] for edge in records_data["edges"]]
        all_records.extend(records)
        
        page_info = records_data["pageInfo"]
        print(f"Página {page_count}: {len(records)} registros")
        
        if not page_info["hasNextPage"]:
            break
            
        cursor = page_info["endCursor"]
        if( page_count == 2):
             break  # Para teste, remova essa linha para obter todos os registros

    print(f"\nTotal de registros obtidos: {len(all_records)}")

   

    # Salvar em arquivo JSON
    with open("pipefy_records.json", "w", encoding="utf-8") as f:
        json.dump(all_records, f, ensure_ascii=False, indent=2)

    # Transformar em tabela para Excel apenas com campos desejados
    campos_desejados = [
        'cnpj',
        'cidade',
        'situação cadastro',
    ]

    campos_desejados_da_busca_cnpj = [
        'uf',
        'cep',
        'email',
        'porte',
        'cnae_fiscal',
        'opcao_pelo_mei',
        'regime_tributario',
        'opcao_pelo_simples',
        'cnae_fiscal_descricao',
        'data_inicio_atividade',
        'data_opcao_pelo_simples',
        'qsa',
        'cnaes_secundarios',  # Lista de sócios
    ]
    # Mapeamento para nomes bonitos no Excel
    campos_nomes_bonitos = {
        'cnpj': 'CNPJ',
        'cidade': 'Cidade',
        'situação cadastro': 'Situação Cadastro',
        'uf': 'UF',
        'cep': 'CEPs',
        'email': 'E-mail',
        'porte': 'Porte',
        'cnae_fiscal': 'CNAE Fiscal',
        'opcao_pelo_mei': 'Opção pelo MEI',
        'regime_tributario': 'Regime Tributário',
        'opcao_pelo_simples': 'Opção pelo Simples',
        'cnae_fiscal_descricao': 'Descrição CNAE Fiscal',
        'data_inicio_atividade': 'Data Início Atividade',
        'data_opcao_pelo_simples': 'Data Opção pelo Simples',
        'qsa': 'Sócios',
        'cnaes_secundarios': 'CNAEs Secundários',
        'telefone': 'Telefone',
    }
    
    records_flat = []
    consulta_realizada = False
    for record in all_records:
        # Verificar se o registro é inativo antes de processar
        situacao_cadastro = None
        if 'record_fields' in record:
            for field in record['record_fields']:
                if field['name'].strip().lower() == 'situação cadastro':
                    situacao_cadastro = field['value']
                    break
        
        # Pular registros inativos
        if situacao_cadastro and 'inativo' in str(situacao_cadastro).lower():
           
            continue
        
        flat = {}
        flat['Cliente'] = record.get('title', '')
        cnpj_valor = None
        if 'record_fields' in record:
            for field in record['record_fields']:
                nome = field['name'].strip().lower()
                if nome in campos_desejados:
                    if nome == 'cnpj':
                        if field['value'] and str(field['value']).strip():
                            flat[field['name']] = colocarMascara_cnpj(field['value'])
                            cnpj_valor = field['value']
                        else:
                            flat[field['name']] = "CNPJ não informado"
                            cnpj_valor = None
                    elif nome == 'cidade':
                        flat[field['name']] = colocarMascara_cidade(field['value'])
                    else:
                        flat[field['name']] = field['value']

        # Consulta telefone e sócios pelo CNPJ para cada registro
        if cnpj_valor:
            todosValores = consulta_cnpj(cnpj_valor, campos_desejados=campos_desejados_da_busca_cnpj)
            if not isinstance(todosValores, dict):
                todosValores = {}
            for campo in campos_desejados_da_busca_cnpj:
                valor = todosValores.get(campo, '')
                if campo == 'qsa':
                    socios = valor if isinstance(valor, list) else []
                    if socios:
                        socios_formatados = []
                        for socio in socios:
                            nome = socio.get('nome_socio', '').strip()
                            qualificacao = socio.get('qualificacao_socio', '').strip()
                            if nome:
                                if qualificacao:
                                    socios_formatados.append(f"{nome} ({qualificacao})")
                                else:
                                    socios_formatados.append(f"{nome}")
                        valor_socios = ', '.join(socios_formatados)
                        if not valor_socios:
                            valor_socios = 'Nenhum sócio encontrado'
                    else:
                        valor_socios = 'Nenhum sócio encontrado'
                    flat[campos_nomes_bonitos.get(campo, campo)] = valor_socios
                elif campo == 'cnaes_secundarios':
                    if isinstance(valor, list):
                        cnaes_formatados = []
                        for cnae in valor:
                            print({cnpj_valor})
                            codigo = str(cnae.get('codigo','')).strip()
                            print({codigo})
                            descricao = cnae.get('descricao','').strip()
                            print({descricao})
                            if codigo and descricao:
                                cnaes_formatados.append(f"{codigo} - {descricao}")
                        valor_cnaes = ','.join(cnaes_formatados)
                    flat[campos_nomes_bonitos.get(campo, campo)] = valor_cnaes
                else:
                    if isinstance(valor, (list, dict)):
                        valor = json.dumps(valor, ensure_ascii=False)
                    flat[campos_nomes_bonitos.get(campo, campo)] = valor
            # Consulta telefone
            telefone = consulta_numero_por_cnpj(cnpj_valor)
            if telefone:
                flat['Telefone'] = telefone
            else:
                flat['Telefone'] = 'Não encontrado'
        else:
            flat['Telefone'] = 'Não encontrado'
            
        for campo in ['cnpj', 'cidade', 'situação cadastro']:
            if campo not in flat:
                flat[campo] = "Não informado"
        records_flat.append(flat)

    
    # Usar nomes bonitos como colunas, se existirem

    colunas_bonitas = ['Cliente'] + ['Telefone'] + ['CNPJ'] + ['Situação Cadastro'] + ['Cidade'] + [campos_nomes_bonitos.get(c, c) for c in campos_desejados_da_busca_cnpj]
    df = pd.DataFrame(records_flat, columns=colunas_bonitas)
    
    # Criar Excel formatado simples
    criar_excel_formatado_com_ordenacao(df, nome_arquivo="pipefy_records_formatado.xlsx", coluna_ordenacao='Cliente', ordem_crescente=True)
    
    print("Arquivo Excel formatado salvo como pipefy_records_formatado.xlsx (apenas campos selecionados)")
    return all_records

# Executar
if __name__ == "__main__":
  get_all_records()